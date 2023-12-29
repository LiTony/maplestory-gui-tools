import numpy
import openpyxl
import random
import functools
import itertools

from PyQt6.QtCore import QDateTime, Qt, QTimer
from PyQt6.QtWidgets import (QApplication, QCheckBox, QComboBox, QDateTimeEdit,
        QDial, QDialog, QGridLayout, QGroupBox, QHBoxLayout, QLabel, QLineEdit,
        QProgressBar, QPushButton, QRadioButton, QScrollBar, QSizePolicy,
        QSlider, QSpinBox, QStyleFactory, QTableWidget, QTabWidget, QTextEdit,
        QVBoxLayout, QWidget, QToolBox, QToolButton)

# >>> c = sheet['B1']
# >>> 'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
# >>> 'Cell ' + c.coordinate + ' is ' + c.value
# >>> sheet['C1'].value

# random.randint(a, b) inclusive both sides

# functools reduce will return 1 value
# itertools accumulate keeps all values

# *args and **kwargs

# debug = True
debug = False
sheet = None

def repeatEGPSimulation(startLevel, numEGP, numRuns):
    finishTotal = 0
    for i in range(numRuns):
        finishTotal += repeatEGP(startLevel, numEGP)
    print("finishTotal: %d" % finishTotal)
    print("finishTotal / numRuns: %.3f" %(finishTotal / numRuns))
    return finishTotal / numRuns

def repeatEGP(startLevel, numEGP):
    level = startLevel
    for i in range(numEGP):
        level = oneEGP(level)
        if(level >= 200):
            return level
    debugPrint("Level Finish: %d", level)
    return level


def oneEGP(currentLevel):
    probValues = getProbability(currentLevel)

    def my_add(a, b):
        return a + b

    # probCumulative = functools.reduce(my_add, probValues, 0)
    probCumulative = list(itertools.accumulate(probValues, my_add))
    debugPrint("Probability Array", probValues)
    debugPrint("Cumulative Array: ", probCumulative)
    randomInteger = random.randint(1, 100)
    debugPrint(randomInteger)

    levelChange = 0

    if(randomInteger <= 50):
        levelChange = startFromLeft(probCumulative, randomInteger)
    else:
        levelChange = startFromRight(probCumulative, randomInteger)
    return currentLevel + levelChange

def startFromLeft(probCumulative, randomInteger):
    length = len(probCumulative)
    i = 0
    while (probCumulative[i] <= randomInteger and i < length):
        i += 1
    debugPrint("Value stopped at: %d index: %d" % (probCumulative[i], i))
    return i+1 #level is index + 1

def startFromRight(probCumulative, randomInteger):
    length = len(probCumulative)
    i = length-1
    while (probCumulative[i] >= randomInteger and i >= 0):
        i -= 1
        if (i == -1):
            debugPrint("probCumulative: ", probCumulative)
            debugPrint("-1, what is randomInteger? %d" % randomInteger)
    debugPrint("Value stopped at: %d index: %d" % (probCumulative[i], i))
    return i+1 #level is index + 1


def getProbability(currentLevel):
    global sheet
    levelRow = currentLevel-128
    row = sheet["A" + str(levelRow)]
    probCellLeft = "B" + str(levelRow)
    probCellRight = "K" + str(levelRow)
    debugPrint("Level: ", row.value)
    probabilities = sheet["%s:%s" % (probCellLeft, probCellRight)][0]
    probabilities = [x.value if x.value != None else 0 for x in probabilities]

    def printProbabilities():
        if(debug):
            print("Probabilities: ")
            for i in range(10):
                prob = probabilities[i]
                print(prob, end=(", " if i!=9 else '\n'))
    printProbabilities()

    return probabilities

def testMain():
    global sheet
    wb = openpyxl.load_workbook("./resources/egp-values.xlsx")
    debugPrint(wb.sheetnames)
    sheet = wb["egp-prob"]
    title = sheet.title
    debugPrint(title)
    activeSheet = wb.active
    debugPrint(activeSheet)

    before = 130
    numPotions = 3
    after = oneEGP(before)
    print("Level Before: %d --- Level After: %d --- Potions Used: %d" \
          % (before, after, 1))
    after = repeatEGP(before, 3)
    print("Level Before: %d --- Level After: %d --- Potions Used: %d" \
          % (before, after, numPotions))
    
    expectedBefore = 141
    expectedPots = 10
    expectedRuns = 1000
    expected = repeatEGPSimulation(expectedBefore, expectedPots, expectedRuns)
    print("Starting at Level %d, with %d potions, expected level is: %d" \
            % (expectedBefore, expectedPots, expected))
# debug / misc code
    
def debugPrint(*args):
    if(debug):
        length = len(args)
        i = 1
        for arg in args:
            print(arg, end=(" " if i!=length else '\n'))
            i += 1

# initialization code
            
def getSheet(fileLocation="./resources/egp-values.xlsx"):
    global sheet
    wb = openpyxl.load_workbook(fileLocation)
    debugPrint(wb.sheetnames)
    sheet = wb["egp-prob"]

class WidgetGallery(QDialog):
    def __init__(self, parent=None):
    
        super(WidgetGallery, self).__init__(parent)

        self.originalPalette = QApplication.palette()
        
        startLevelLabel = QLabel("&Level:")
        startLevel = QSpinBox()
        startLevel.setMinimum(130)
        startLevel.setMaximum(199)
        startLevel.setObjectName("startLevel")
        startLevel.setReadOnly(False)
        startLevelLabel.setBuddy(startLevel)

        numberPotionsLabel = QLabel("&Number of Potions:")
        numberPotions = QSpinBox()
        numberPotions.setMinimum(0)
        numberPotions.setMaximum(100)
        numberPotions.setObjectName("potions")
        numberPotions.setReadOnly(False)
        numberPotionsLabel.setBuddy(numberPotions)

        runsToPerformLabel = QLabel("&Number of Simulations")
        runsToPerform = QSpinBox()
        runsToPerform.setMinimum(100)
        runsToPerform.setMaximum(10000)
        runsToPerform.setObjectName("simulations")
        runsToPerformLabel.setBuddy(runsToPerform)

        topLayout = QGridLayout()
        topLayout.addWidget(startLevelLabel)
        topLayout.addWidget(startLevel)
        topLayout.addWidget(numberPotionsLabel)
        topLayout.addWidget(numberPotions)
        topLayout.addWidget(runsToPerformLabel)
        topLayout.addWidget(runsToPerform)



        expectedLevelLabel = QLabel("&Expected Level:")
        expectedLevel = QLineEdit()
        expectedLevel.setReadOnly(True)
        expectedLevel.setObjectName("expectedLevel")
        expectedLevelLabel.setBuddy(expectedLevel)

        bottomLayout = QGridLayout()
        bottomLayout.addWidget(expectedLevelLabel)
        bottomLayout.addWidget(expectedLevel)

        self.createRunBar()
        self.createResetButton()

        mainLayout = QGridLayout()
        mainLayout.addLayout(topLayout, 0, 0, 1, 4)
        mainLayout.addLayout(bottomLayout, 1, 0, 1, 4)
        mainLayout.addWidget(self.runBar, 2, 0, 1, 3)
        mainLayout.addWidget(self.resetButton, 2, 3)

        self.runBar.findChild(QPushButton) \
            .clicked.connect(self.runPress)
        self.resetButton.findChild(QPushButton) \
            .clicked.connect(self.resetPress)
        self.setLayout(mainLayout)

        self.setWindowTitle("EGP Simulator App by LiTony")

        getSheet()

    def createRunBar(self):
        self.runBar = QGroupBox("Run")

        layout = QVBoxLayout()
        button = QPushButton('Run')
        button.setDefault(False)
        button.setObjectName('RunBar')
        layout.addWidget(button)
        
        self.runBar.setLayout(layout)

    def createResetButton(self):
        self.resetButton = QGroupBox("Reset")

        layout = QVBoxLayout()
        button = QPushButton('Reset')
        button.setDefault(False)
        button.setObjectName('ResetButton')
        layout.addWidget(button)
        
        self.resetButton.setLayout(layout)

    def runPress(self):
        potionCount = self.findChild(QSpinBox, "potions").value()
        startLevel = self.findChild(QSpinBox, "startLevel").value()
        simulations = self.findChild(QSpinBox, "simulations").value()
        result = repeatEGPSimulation(startLevel, potionCount, simulations)
        self.findChild(QLineEdit, "expectedLevel").setText(str(result))

    def resetPress(self):
        self.findChild(QSpinBox, "startLevel").setValue(141)
        self.findChild(QSpinBox, "potions").setValue(0)
        self.findChild(QLineEdit, "expectedLevel").setText("")
        self.findChild(QSpinBox, "simulations").setValue(100)
        print("input fields reset")

if __name__ == '__main__':

    import sys

    app = QApplication(sys.argv)
    gallery = WidgetGallery()
    gallery.show()
    sys.exit(app.exec())